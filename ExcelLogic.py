import glob

from openpyxl import *
import os
from datetime import datetime

class excel():

    def GetExcelFile(self, final: str, reportname,excelName = "ExcelFebDev"):
        #print(os.path.exists("ExcelNew.xlsx"))
        ExcelFileName = excelName+".xlsx"
        if os.path.exists(ExcelFileName):
            wb = load_workbook(ExcelFileName)
            ws = wb.active
            data = [(reportname, final, 'NA', 'NA')]
            for i in data:
                ws.append(i)
        else:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Report Name"
            ws["B1"] = "Path of Generated reports"
            ws["C1"] = "Path of Standard reports"
            ws["D1"] = "Result"
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 10
            data = [(reportname, final, 'NA', 'NA')]
            for j in data:
                ws.append(j)
        wb.save(ExcelFileName)

    def infoLog(self,listoffailedfiles):
        data = []
        InfoExcelFileName = f"infolog_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
        if os.path.exists(InfoExcelFileName):
            wb = load_workbook(InfoExcelFileName)
            ws = wb.active
            for reason in listoffailedfiles:
                data = [(reason[0], reason[1], reason[2], 'Fail')]
                for i in data:
                    ws.append(i)
        else:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Path of report"
            ws["B1"] = "Name of report"
            ws["C1"] = "Reason to fail"
            ws["D1"] = "Result"
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 25
            for reason in listoffailedfiles:
                data = [(reason[0], reason[1], reason[2], 'Fail')]
                for j in data:
                    ws.append(j)
        wb.save(InfoExcelFileName)

    def openExcel(self):
        for excelfile in glob.glob("*.xlsx"):
            os.system(f"start EXCEL.EXE {excelfile}")


